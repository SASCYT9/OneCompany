#!/usr/bin/env python3
"""Parse the 2026 April V2.0 iPE price list xlsx into IpeParsedPriceList JSON.

Structure: each sheet contains alternating header rows + brand banners +
table headers (MODEL, SECTION, SKU NO., MAT., DESCRIPTION, ..., MSRP(USD))
followed by data rows. Model column is multi-line "BRAND\\nMODEL CHASSIS\\nYEARS".
Section column carries section. SKU + Mat + Desc + Price across the row.

Same retail-pricing formula as the existing PDF parser: flat fee by threshold.

Usage:
  python scripts/parse-ipe-xlsx-pricelist.py scripts/ipe/2026-april-pricelist.xlsx \\
    --json-out artifacts/ipe-price-list/2026-04-pricelist.parsed.json
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Iterable

import openpyxl


BRAND_ALIASES: dict[str, str] = {
    "AUDI": "Audi",
    "Audi": "Audi",
    "Aston Martin": "Aston Martin",
    "BMW": "BMW",
    "Benz": "Mercedes-Benz",
    "Mercedes": "Mercedes-Benz",
    "Mercedes-Benz": "Mercedes-Benz",
    "Mercedes Benz": "Mercedes-Benz",
    "Bentley": "Bentley",
    "Chevrolet": "Chevrolet",
    "Ferrari": "Ferrari",
    "Jaguar": "Jaguar",
    "Lamborghini": "Lamborghini",
    "Land Rover": "Land Rover",
    "Lexus": "Lexus",
    "MINI": "MINI",
    "Maserati": "Maserati",
    "Mclaren": "McLaren",
    "McLaren": "McLaren",
    "Nissan": "Nissan",
    "Porsche": "Porsche",
    "Rolls Royce": "Rolls-Royce",
    "Ford": "Ford",
    "Subaru": "Subaru",
    "Toyota": "Toyota",
    "Volkswagen": "Volkswagen",
    "VW": "Volkswagen",
    "Honda": "Honda",
    "Hyundai": "Hyundai",
}
NORMALIZED_BRAND_ALIASES: dict[str, str] = {}
for alias, brand in BRAND_ALIASES.items():
    NORMALIZED_BRAND_ALIASES[alias.strip().lower()] = brand
    NORMALIZED_BRAND_ALIASES[brand.strip().lower()] = brand

YEAR_RE = re.compile(r"((?:19|20)\d{2}\s*[-–]\s*(?:Current|(?:19|20)\d{2}))", re.IGNORECASE)
ENGINE_RE = re.compile(r"\bengine\b|\bturbo\b|\bv\d\b|\bw\d\b|\b\d\.\dT?\b", re.IGNORECASE)
TBD_RE = re.compile(r"\bTBD\b", re.IGNORECASE)


def collapse(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("：", ":")).strip()


def resolve_brand(value: str) -> str | None:
    return NORMALIZED_BRAND_ALIASES.get(value.strip().lower())


def looks_like_table_header(row: list[Any]) -> bool:
    norm = [collapse(c).upper() for c in row]
    return "MODEL" in norm and "SKU NO." in norm and any("MSRP" in c for c in norm)


def parse_money(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def split_model_cell(cell: str) -> tuple[str | None, str | None, str | None]:
    """Split "BRAND\\nMODEL CHASSIS ENGINE\\nYEAR-YEAR" into (model, year, engine).

    BRAND is consumed by the section banner. We try to keep model + chassis +
    engine as a single descriptor, then peel out the year-range and any obvious
    engine token for the structured fields the import logic expects.
    """
    if not cell:
        return None, None, None
    lines = [collapse(line) for line in cell.splitlines() if collapse(line)]
    # First line is brand (e.g., "AUDI"). Drop it if it matches a known brand.
    if lines and resolve_brand(lines[0]):
        lines = lines[1:]
    if not lines:
        return None, None, None

    year_range = None
    engine = None
    model_lines: list[str] = []
    for line in lines:
        m = YEAR_RE.search(line)
        if m and not year_range:
            year_range = collapse(m.group(1))
            cleaned = YEAR_RE.sub("", line).strip(" -–")
            if cleaned:
                model_lines.append(cleaned)
            continue
        model_lines.append(line)

    model = " | ".join(model_lines) if model_lines else None
    if model:
        em = ENGINE_RE.search(model)
        if em:
            engine = em.group(0)
    return model, year_range, engine


def is_blank(row: Iterable[Any]) -> bool:
    return all(cell is None or collapse(cell) == "" for cell in row)


def detect_columns(header_row: list[Any]) -> dict[str, int]:
    cols: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        norm = collapse(cell).upper()
        if not norm:
            continue
        if norm == "MODEL" and "model" not in cols:
            cols["model"] = idx
        elif norm == "SECTION":
            cols["section"] = idx
        elif "SKU" in norm:
            cols["sku"] = idx
        elif norm in {"MAT.", "MATERIAL"}:
            cols["material"] = idx
        elif norm == "DESCRIPTION":
            cols["description"] = idx
        elif "MSRP" in norm:
            cols["msrp"] = idx
    return cols


def compute_import_fee(msrp: float | None, low_fee: float, high_fee: float, threshold: float) -> float | None:
    if msrp is None:
        return None
    return high_fee if msrp >= threshold else low_fee


def parse_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    tier_label: str | None,
    low_fee: float,
    high_fee: float,
    threshold: float,
) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    items: list[dict[str, Any]] = []

    current_brand: str | None = None
    columns: dict[str, int] = {}
    last_model: str | None = None
    last_year: str | None = None
    last_engine: str | None = None
    last_section: str | None = None
    pending_remarks: list[str] = []

    in_remarks = False
    page_no = 1  # synthetic — entire sheet is page 1 to keep schema parity

    for line_index, raw_row in enumerate(rows):
        row = list(raw_row) + [None] * (max(0, 21 - len(raw_row)))
        first = collapse(row[0])
        if not first and is_blank(row):
            continue

        # Brand banner: row with just a brand name in col A
        if first and not looks_like_table_header(row):
            non_empty_count = sum(1 for c in row if collapse(c))
            if non_empty_count == 1 and resolve_brand(first):
                current_brand = resolve_brand(first)
                in_remarks = False
                continue

        # Table header — refresh column map
        if looks_like_table_header(row):
            columns = detect_columns(row)
            in_remarks = False
            continue

        # Remark blocks
        if first == "Remark" or first.startswith("Remark"):
            in_remarks = True
            continue

        if first.startswith("※") or in_remarks:
            joined = " | ".join(collapse(c) for c in row if collapse(c))
            if joined:
                pending_remarks.append(joined)
            in_remarks = True
            continue

        if not columns:
            continue

        col = columns
        sku_val = collapse(row[col.get("sku", 2)] if "sku" in col else None)
        if not sku_val:
            # Possibly model row only — pull the model context for upcoming SKUs
            model_cell = row[col["model"]] if "model" in col else None
            section_cell = row[col["section"]] if "section" in col else None
            if model_cell:
                m, y, e = split_model_cell(str(model_cell))
                if m:
                    last_model = m
                if y:
                    last_year = y
                if e:
                    last_engine = e
            if section_cell and collapse(section_cell):
                last_section = collapse(section_cell)
            continue

        # Pull model/section context — model_cell may be empty when SKU continues
        # the previous section.
        model_cell = row[col["model"]] if "model" in col else None
        if model_cell and collapse(model_cell):
            m, y, e = split_model_cell(str(model_cell))
            if m:
                last_model = m
            if y:
                last_year = y
            if e:
                last_engine = e
            # New model row resets remarks
            pending_remarks = []
        section_cell = row[col["section"]] if "section" in col else None
        if section_cell and collapse(section_cell):
            last_section = collapse(section_cell)

        material = collapse(row[col["material"]] if "material" in col else None) or None
        description = collapse(row[col["description"]] if "description" in col else None)

        # Header places "MSRP(USD)" in col G but the actual value sits in col H
        # (and col G is reserved for a "+" relative-price marker). Look one
        # column to the right of the header position for the numeric MSRP.
        msrp_idx = col.get("msrp")
        msrp = None
        rel_marker = ""
        if msrp_idx is not None:
            for probe in (msrp_idx + 1, msrp_idx):
                if probe < len(row):
                    candidate = parse_money(row[probe])
                    if candidate is not None:
                        msrp = candidate
                        # The "+" marker (when present) lives in the cell
                        # immediately before the value column.
                        rel_marker = collapse(row[probe - 1]) if probe - 1 >= 0 else ""
                        break

        # The 2026 sheet uses an extra column F as a contextual descriptor.
        # Stitch it into description if helpful.
        extra = collapse(row[col["description"] + 1]) if "description" in col and col["description"] + 1 < len(row) else ""
        if extra and extra not in description:
            description = collapse(f"{description} | {extra}") if description else extra

        price_kind: str | None = None
        if msrp is not None:
            price_kind = "relative" if rel_marker == "+" else "absolute"
        elif any(TBD_RE.search(collapse(c)) for c in row):
            price_kind = "tbd"

        import_fee = compute_import_fee(msrp, low_fee, high_fee, threshold) if price_kind == "absolute" else None
        retail = (msrp + import_fee) if (msrp is not None and import_fee is not None) else (msrp if price_kind in {"relative", "free", "included"} else None)

        items.append(
            {
                "page": page_no,
                "tier": tier_label,
                "brand": current_brand or "Unknown",
                "model": last_model,
                "year_range": last_year,
                "engine": last_engine,
                "section": last_section,
                "sku": sku_val,
                "material": material,
                "description": description,
                "price_kind": price_kind or "unpriced",
                "msrp_usd": msrp,
                "import_fee_usd": import_fee,
                "retail_usd": retail,
                "remarks": " | ".join(pending_remarks) if pending_remarks else None,
                "matched_summary_label": None,
                "raw_line": " | ".join(collapse(c) for c in row if collapse(c)),
            }
        )

    return items


def parse_addons_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, Any]:
    """Parse the 'Add-on options' sheet into structured tip/valve/accessory lists.

    Column layout (0-indexed):
      0: section label + sub-label joined by '\\n' (only on first row of each section)
      2: SKU (Valve Controls only)
      3: MAT (SS / Ti)
      4: DESCRIPTION (e.g. "Chrome/Polished Silver Tips")
      6: relative-price marker '+' (Valve Controls + Accessories)
      7: MSRP price (USD)
    """
    rows = list(ws.iter_rows(values_only=True))

    section_label: str | None = None
    dual_tips: list[dict[str, Any]] = []
    quad_tips: list[dict[str, Any]] = []
    valve_controls: list[dict[str, Any]] = []
    accessories: list[dict[str, Any]] = []
    in_remarks = False

    SECTION_KEYS = ("Valve Controls", "Dual Out Tips", "Quad Tips", "Accessories")

    for raw_row in rows:
        row = list(raw_row) + [None] * (max(0, 9 - len(raw_row)))
        first_raw = row[0] if row[0] is not None else ""
        first = collapse(first_raw)
        # Section label often arrives as "Section\nsub-label" — match on the
        # first line only.
        section_first_line = str(first_raw).splitlines()[0].strip() if first_raw else ""

        if first == "Remark" or first.startswith("Remark") or first.startswith("Notice"):
            in_remarks = True
            continue
        if in_remarks:
            continue
        if first.startswith("※"):
            continue
        if not first and is_blank(row):
            continue
        if looks_like_table_header(row):
            continue

        if section_first_line in SECTION_KEYS:
            section_label = section_first_line

        if section_label is None:
            continue

        # Price at col 7
        price = parse_money(row[7]) if row[7] is not None else None
        if price is None:
            continue

        sku = collapse(row[2]) or None
        material = collapse(row[3]) or None
        # Description may have a "\n" continuation (Titanium Material rows).
        description = collapse(row[4]) or ""
        description = description.replace("\n", " — ") if description else ""

        entry = {
            "label": description or "",
            "material": material,
            "msrp_usd": price,
            "sku": sku,
        }

        if section_label == "Valve Controls":
            valve_controls.append(entry)
        elif section_label == "Dual Out Tips":
            dual_tips.append(entry)
        elif section_label == "Quad Tips":
            quad_tips.append(entry)
        elif section_label == "Accessories":
            accessories.append(entry)

    for collection in (dual_tips, quad_tips):
        if not collection:
            continue
        cheapest = min(collection, key=lambda e: e["msrp_usd"])
        for e in collection:
            e["is_default"] = e is cheapest

    return {
        "dual_tips": dual_tips,
        "quad_tips": quad_tips,
        "valve_controls": valve_controls,
        "accessories": accessories,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Parse iPE 2026 xlsx price list.")
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument("--json-out", type=Path, required=True)
    parser.add_argument("--addons-json-out", type=Path, default=None)
    parser.add_argument("--low-fee", type=float, default=1500)
    parser.add_argument("--high-fee", type=float, default=1600)
    parser.add_argument("--threshold", type=float, default=5000)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    items: list[dict[str, Any]] = []
    if "2026 Premium products" in wb.sheetnames:
        items.extend(
            parse_sheet(
                wb["2026 Premium products"],
                tier_label="Premium Products",
                low_fee=args.low_fee,
                high_fee=args.high_fee,
                threshold=args.threshold,
            )
        )
    if "Standard products" in wb.sheetnames:
        items.extend(
            parse_sheet(
                wb["Standard products"],
                tier_label="Standard Products",
                low_fee=args.low_fee,
                high_fee=args.high_fee,
                threshold=args.threshold,
            )
        )

    args.json_out.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source_pdf": str(args.xlsx_path),
        "pricing_formula": {
            "type": "flat_import_fee_by_threshold",
            "threshold_usd": args.threshold,
            "low_fee_usd": args.low_fee,
            "high_fee_usd": args.high_fee,
            "rule": f"retail = msrp + ({args.high_fee} if msrp >= {args.threshold} else {args.low_fee})",
        },
        "items": items,
    }
    args.json_out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    priced = sum(1 for it in items if it["msrp_usd"] is not None)
    retail_ready = sum(1 for it in items if it["retail_usd"] is not None)
    by_brand: dict[str, int] = {}
    for it in items:
        by_brand[it["brand"]] = by_brand.get(it["brand"], 0) + 1
    print(f"Total rows: {len(items)}")
    print(f"Priced: {priced}")
    print(f"Retail-ready: {retail_ready}")
    print(f"By brand: {sorted(by_brand.items(), key=lambda x: -x[1])[:15]}")
    print(f"JSON: {args.json_out}")

    # Addons sheet (Tips / Valves / Accessories) — emits a separate JSON when
    # --addons-json-out is provided.
    if args.addons_json_out is not None and "Add-on options" in wb.sheetnames:
        addons = parse_addons_sheet(wb["Add-on options"])
        args.addons_json_out.parent.mkdir(parents=True, exist_ok=True)
        addons_payload = {
            "source_xlsx": str(args.xlsx_path),
            **addons,
        }
        args.addons_json_out.write_text(
            json.dumps(addons_payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        print(
            f"Addons: dual_tips={len(addons['dual_tips'])} quad_tips={len(addons['quad_tips'])} "
            f"valves={len(addons['valve_controls'])} accessories={len(addons['accessories'])}"
        )
        print(f"Addons JSON: {args.addons_json_out}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
